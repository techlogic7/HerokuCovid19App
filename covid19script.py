#Importing packages
import sys
import os
import schedule
import pandas as pd
import numpy as np
import requests
import time
from bs4 import BeautifulSoup, NavigableString, Tag
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium import webdriver
from datetime import datetime,date
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
import pandas as pd
from openpyxl import load_workbook



# sys.exit()

print("sys.path",sys.path)
print("sys.executable",sys.executable)
print("os.getcwd()",os.getcwd())
print("os.path.abspath(__file__)",os.path.abspath(__file__))
# sys.exit()
def scrape():
    driver =  False 
    try:
        print("Selenium chrome driver is set to launch")
        chrome_options = webdriver.ChromeOptions()
        chrome_options.binary_location = os.environ.get("GOOGLE_CHROME_BIN")
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--no-sandbox")
        driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), chrome_options=chrome_options)

        # options = Options()
        # options.headless = True
        # driver = webdriver.Chrome(ChromeDriverManager().install(),options=options) ##instace of selenium driver

        url = "https://www.mohfw.gov.in/"
        driver.get(url)  #we are getting data here
        TIMEOUT  = 10
        WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div/section[3]/div/div/div/div/table/tbody"))
        )
        webcontent = driver.page_source
        PySoup = BeautifulSoup(webcontent,'html.parser')
        print("Loading is complete!")
    except TimeoutException:
        print("Loading took too much time!")
    finally:
        if driver:
            driver.quit()
    # sys.exit()
    divdata = PySoup.find('div',{'class':'data-table table-responsive'})
    divdate = divdata.h5.span.text.split(",")[0].split(":")[1].strip()
    format = "%d %B %Y"
    datetimeobj = datetime.strptime(divdate,format).date()


    # table_name = statetable table table-striped or you can directly retrive tbody from the beautifulSoup Object
    mylistoftr = PySoup.find('tbody').find_all('tr') #it will contain state data

    mytabledata  = []
    headings =  ["Sr.No","States/UT","Active Cases","Active Cases Since Yesterday",\
                "Recovered Cases","Recovered Cases Since Yesterday","Deceased Cases","Deceased Cases Since Yesterday"]
    for tr in mylistoftr:
        if isinstance(tr,Tag):
            statedata = [-int(td.text.strip()) \
                        if (td.span != None) and (td.span.get('class') in [['down']]) \
                        else td.text.strip()
                        for td in tr.find_all('td')]
            if 'Total#' in statedata:
                statedata.insert(1,"")
            data = dict(zip(headings,statedata))
            mytabledata.append(data)
    covid19df =pd.DataFrame(mytabledata,columns=headings)

    if not covid19df.index.name =='Sr.No':
        covid19df.set_index('Sr.No',inplace=True)
    covid19df = covid19df[0:35]
    covid19df.replace("", 0,inplace=True)
    covid19df.replace(np.nan,0,inplace=True)
    # state_data.replace(['',np.nan], 0,inplace=True)
    covid19df['Date'] = datetimeobj.strftime("%d-%m-%Y")
    covid19df.rename(index={'Total#': 'Total'},inplace=True)
    cols=[colname for colname in covid19df.columns if colname not in ['Total','States/UT','Date']]
    covid19df[cols] = covid19df[cols].astype('int')

    if len(covid19df) == 35:
        return [True,covid19df]
    else:
        return [False,covid19df]



def append_df_to_excel(filename,df,sheetname="newsheet",startrow=None,startcol = None,appendby=None,truncate_sheet=False, **to_excel_kwargs):
    
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      startcol : right most cell columns to dump data frame.
                 Per default (startcol=None) calculate the last column
                 in the existing DF and write to the next column...
                 
      append_by : takes string type of either "row" or "column" ...
                  default: "row"
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
        
    writer = pd.ExcelWriter(filename, engine='openpyxl')   #write_obj for the excel sheet
    # try to open an existing workbook
    last_col_val= False
    first_col_val = False
    try:
        check = open(filename,'r')
        check.close()
        book = load_workbook(filename)
        writer.book = book
        
        sheet_obj = book.active 
        max_column = sheet_obj.max_column     
        max_row = sheet_obj.max_row


        last_col_val_1strow = sheet_obj.cell(row = 1, column = max_column).value
        last_col_val_lastrow = sheet_obj.cell(row = max_row, column = max_column).value
        datestr = df['Date'][0]
        format = "%d-%m-%Y"
        datetimeobj = datetime.strptime(datestr,format).date()
        if last_col_val_1strow == datetimeobj.strftime("%d-%m-%Y"):
            print("{} is already updated for {}".format(filename,datetimeobj.strftime("%d-%m-%Y")))
            return True
        if last_col_val_lastrow == datetimeobj.strftime("%d-%m-%Y"):
            print("{} is already updated for {}".format(filename,datetimeobj.strftime("%d-%m-%Y")))
            return True
        first_col_val = sheet_obj.cell(row = 1, column = 1).value
        if 'mycovid19' in filename:
            appendby = "row"
        if 'mycovid19' in filename and sheetname in writer.book.sheetnames and appendby == "row":
            startrow = writer.book[sheetname].max_row
        elif not startrow:
            startrow  = 0
   
        if startcol is None and sheetname in writer.book.sheetnames and appendby == "column":
            startcol = writer.book[sheetname].max_column
        elif not startcol:
            startcol = False
            
        if not startrow and not startcol and sheetname not in writer.book.sheetnames:
            to_excel_kwargs['header'] = True

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheetname)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheetname, idx)

        # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        #read existing file
        reader = pd.read_excel(filename)
    except FileNotFoundError:
        print("{} file not found, so creating a one".format(filename))
        # file does not exist yet, we will create it
        if not startrow:
            startrow = False
        if not startcol:
            startcol = False
        if not startrow and not startcol and not to_excel_kwargs['header']:
            to_excel_kwargs['header'] = True
        pass
    except Exception as e:
        if 'Permission denied' in str(e):
            print("Error : Permission denied {} is currently opened in your Excel , Please close and try again!!".format(filename))
        else:    
            print(e)
        return False
    
        # write out the new sheet
        
# # ======>>> for active /Confirmed cases <<<===========

    if 'active_cases' in filename:
        df[datetimeobj.strftime("%d-%m-%Y")] = df['Active Cases'].values
    if 'recovered_cases' in filename:
        df[datetimeobj.strftime("%d-%m-%Y")] = df['Recovered Cases'].values
    if 'deceased_cases' in filename:
        df[datetimeobj.strftime("%d-%m-%Y")] = df['Deceased Cases'].values
        
    if datetimeobj.strftime("%d-%m-%Y") in df.columns:
        if not first_col_val and first_col_val != "States/UT":
            print("{} created to date {}".format(filename,datetimeobj.strftime("%d-%m-%Y")))
            df[['States/UT',datetimeobj.strftime("%d-%m-%Y")]].to_excel(writer,sheet_name=sheetname,startrow=startrow,startcol= startcol,**to_excel_kwargs)
        else:
            print("{} updated to date {}".format(filename,datetimeobj.strftime("%d-%m-%Y")))
            df[datetimeobj.strftime("%d-%m-%Y")].to_excel(writer,sheet_name=sheetname,startrow=startrow,startcol= startcol,**to_excel_kwargs)

        df.drop([datetimeobj.strftime("%d-%m-%Y")], axis=1,inplace = True)
        writer.close()
    if 'mycovid19' in filename and datetimeobj.strftime("%d-%m-%Y") not in df.columns:
        print("{} updated to date {}".format(filename,datetimeobj.strftime("%d-%m-%Y")))
        df.to_excel(writer,sheet_name=sheetname,startrow=startrow,startcol = startcol,**to_excel_kwargs)
        writer.close()



def updater():
    checkdf = scrape()
    if checkdf[0]:
        covid19df = checkdf[1]
        activefilename  = r'active_cases.xlsx'
        recoveredfilename  = r'recovered_cases.xlsx'
        deceasedfilename  = r'deceased_cases.xlsx'
        covid19  = r'mycovid19.xlsx'
        myfiles  =  [activefilename,covid19,recoveredfilename,deceasedfilename]
        for file in myfiles:
            mysheet = "COVID19_TIMESERIESDATA"
            append_df_to_excel(file,covid19df,sheetname=mysheet,startrow=None,startcol=None,appendby="column",
                        truncate_sheet=False,header = True,index=False)
        print("Script executed successfully !!")
    # break


# updater()

# Task scheduling 

# After every 10mins updater() is called.  
schedule.every(2).minutes.do(updater)
  
# After every hour updater() is called. 
# schedule.every().hour.do(updater)
  
# Every day at 12am or 00:00 time updater() is called. 
# schedule.every().day.at("10:41").do(updater)
# schedule.every().day.at("00:00").do(updater)

while True:
    schedule.run_pending()
    time.sleep(1) # wait one minute



# def custom_strftime(format,date_obj):
#     suffix = (lambda X:'th' if 11<=X<=13 else {1:'st',2:'nd',3:'rd'}.get(X%10, 'th'))(int(date_obj.day))
#     return date_obj.strftime(format).replace('{today_day}', str(date_obj.day) + str(suffix))
# file_name = "COVID19_{}".format(custom_strftime('{today_day}%b',datetimeobj))
# covid19df.to_excel( r'{}.xlsx'.format(file_name),sheet_name = 'COVID19 State Data')
# covid19df.to_csv( r'{}.csv'.format(file_name))